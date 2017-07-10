"""
This program will run on monthly basis and collect ASP Customers data from ASP EA Data Base.
The objective is to collect ASP Customer data for GE Billing team to be able to create invoice.
Data will be compiled into Excel worksheet and emailed to Fiance team.
"""
import pyodbc
import datetime
import smtplib
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders


def ea_monthly_report():

    """
    This function will do 4 tasks:
    1.  Connect to SQL Server.
    2.  Run query and retrieve data from SQL server Data Base.
    3.  Write the data to Excel worksheet.
    4.  Email the worksheet to end user.
    """

    # File name for Excel Worksheet
    excel_filename = ""

    # Get all Virtual Archive names from SQL Server
    virtual_archives = get_archives()

    # Set up Excel Worksheet.
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "EA Monthly Report"
    work_sheet.append(['Archive Name', 'Study Month-Year', 'Total Exams', 'Total Size in MB', 'Average Exam Size in MB', 'Total Size in GB', 'Average Exam Size in GB'])

    # Assign font and background color properties for Column Title cells
    f = Font(name="Arial", size=14, bold=True, color="FF000000")
    fill = PatternFill(fill_type="solid", start_color="00FFFF00")

    work_sheet["A1"].fill = fill
    work_sheet["B1"].fill = fill
    work_sheet["C1"].fill = fill
    work_sheet["D1"].fill = fill
    work_sheet["E1"].fill = fill
    work_sheet["F1"].fill = fill
    work_sheet["G1"].fill = fill

    work_sheet["A1"].font = f
    work_sheet["B1"].font = f
    work_sheet["C1"].font = f
    work_sheet["D1"].font = f
    work_sheet["E1"].font = f
    work_sheet["F1"].font = f
    work_sheet["G1"].font = f

    # Set Column width
    work_sheet.column_dimensions["A"].width = 25.0
    work_sheet.column_dimensions["B"].width = 25.0
    work_sheet.column_dimensions["C"].width = 25.0
    work_sheet.column_dimensions["D"].width = 25.0
    work_sheet.column_dimensions["E"].width = 40.0
    work_sheet.column_dimensions["F"].width = 40.0
    work_sheet.column_dimensions["G"].width = 40.0

    # Obtain Exam Volume for all virtual archives and write the data to excel sheet.
    for archive in virtual_archives:
        # Obtains archive volume from SQL server.
        rows = archive_volume(archive)
        for row in rows:
            average_exam_volume_in_gb = round((average_exam_size(row[1], row[2]) / 1024), 6)
            # Adds archive name and exam volume to Worksheet.
            work_sheet.append([archive,                             # Archive Name
                               format_date(row[0]),                 # Study By Year
                               int(row[1]),                         # Total Exams
                               round(row[2], 2),                    # Total Size in MB
                               average_exam_size(row[1], row[2]),   # Average Exam Size in MB
                               exam_size_in_gb(row[2]),             # Total Size in GB
                               average_exam_volume_in_gb            # Average Exam Size in GB
                               ])
            # Format cells to use 1000 comma separator.
            work_sheet['C{}'.format(work_sheet.max_row)].style = 'Comma [0]'
            work_sheet['D{}'.format(work_sheet.max_row)].style = 'Comma'
            work_sheet['E{}'.format(work_sheet.max_row)].style = 'Comma'
            work_sheet['F{}'.format(work_sheet.max_row)].style = 'Comma'

        print("Added {} Exam Volume to Workbook!".format(archive))
        # Add blank line
        work_sheet.append([])

    # Saves Excel worksheet.
    excel_filename = "ASP_Monthly_Report_{}.xlsx".format(datetime.datetime.now().strftime("%Y-%m-%d"))
    work_book.save(excel_filename)

    # Send email with attachment.
    send_email(excel_filename)


# Obtain Archive Names from SQL Server.
def get_archives():
    """
    This function will return list of Virtual Archive Names from SQL Server.
    """
    # Create list to store archive names.
    archive_list = []

    # Define data base connection parameters.
    sqlserver = 'SQL1'
    database = 'RSAdmin'
    username = 'admin'
    password = 'Bosnia66s'

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER=' + sqlserver + ';'
        r'DATABASE=' + database + ';'
        r'UID=' + username + ';'
        r'PWD=' + password + ''
    )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                SELECT DBName from tblArchive
                ORDER BY DBName
                """)
    rows = cur.fetchall()

    # Add Archive names to archive list.
    for row in rows:
        archive_list.append(row[0])
    # Close SQL Connection.
    cur.close()
    conn.close()

    return archive_list


# Calculate last day of month
def last_day_of_month(any_day):
    """This function will return last day of the month.
    """
    current_month = any_day.replace(day=1)
    return current_month - datetime.timedelta(days=current_month.day)


# Calculate first day of month
def first_day_of_month(any_day):
    """This function will return first day of month. For the parameter it requires
     output from last_day_of_month() function.
     """
    return any_day.replace(day=1)


# Obtain archive volume.
def archive_volume(db_name):
    """
    This function will obtain archive volume form SQL server.
    """
    # Calculate first and last day of month.
    today = datetime.date.today()
    last_day = last_day_of_month(today)
    first_and_last_month_day = [str(first_day_of_month(last_day)), str(last_day)]

    # Define data base connection parameters.
    sqlserver = 'SQL1'
    database = db_name
    username = 'admin'
    password = 'Bosnia66s'

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER='+sqlserver+';'
        r'DATABASE='+database+';'
        r'UID='+username+';'
        r'PWD='+password+''
        )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                select convert(varchar(6), FirstArchiveDate, 112) as StudyYear, count(distinct id1) as TotalExam, sum(t.ByteSize/(1024*1024)) AS StudySizeMB
                from tblDICOMStudy s with (nolock), tblDICOMSeries se with (nolock), tblDICOMImage im with (nolock), tblFile t with (nolock) 
                where s.Id1 = se._Id1 and se.Id2 = im._Id2 and im._idFile = t.idFile
                group by convert(varchar(6), FirstArchiveDate, 112)
                order by convert(varchar(6), FirstArchiveDate, 112)
                """)
    rows = cur.fetchall()
    # Close SQL Server Connections.
    cur.close()
    conn.close()
    return rows


# Send email with Report
def send_email(file_attachment):
    """This function will send email with the attachment.
    It takes attachment file name as argument.
    """

    # Define email body
    body = "This is EA Monthly report. See attached file for Total Exam Volume for each customer."
    content = MIMEText(body, 'plain')

    # Open file attachment
    filename = file_attachment
    infile = open(filename, "rb")

    # Set up attachment to be send in email
    part = MIMEBase("application", "octet-stream")
    part.set_payload(infile.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)

    msg = MIMEMultipart("alternative")

    # Define email recipients
    to_email = "nerminkekic@ge.com"
    from_email = "nerminkekic@ge.com"

    # Create email content
    msg["Subject"] = "ASP Monthly Report {}".format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    msg["From"] = from_email
    msg["To"] = to_email
    msg.attach(part)
    msg.attach(content)

    # Send email to SMTP server
    s = smtplib.SMTP("10.4.1.1", 25)
    s.sendmail(from_email, to_email, msg.as_string())
    s.close()


# Perform calculation for Average Exam size in MB
def average_exam_size(exam_volume, total_storage_size):
    """Calculate average exam size in MB"""
    return round((total_storage_size / exam_volume), 2)


# Convert from MB to GB
def exam_size_in_gb(size_in_mb):
    """Convert Average exam size in MB to GB"""
    return round((size_in_mb / 1024), 2)


# Format Data Base time to use format MM-YYY
def format_date(db_date):
    """Format Data Base time to use format MM-YYY
    """
    try:
        dt = datetime.date(int(db_date[0:4]), int(db_date[4:6]), 1)
    except TypeError:
        print("There is no valid Study Date in Data base!")
    except Exception as e:
        print(e)
    else:
        return dt.strftime("%b{}%y".format("-"))


# Run script
ea_monthly_report()
